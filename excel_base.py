from io import BytesIO
from typing import Optional

from openpyxl import load_workbook

import database


# Columnas de datos que se escriben/borran; el resto (C–F, G con fórmula) no se toca.
_DATA_COLS = (1, 2, 8, 9)   # A, B, H, I  (índices 1-based de openpyxl)


# ── Helpers ───────────────────────────────────────────────────────────────────

def _load_excel_bytes(project_name: str) -> Optional[bytes]:
    """Return raw Excel bytes from DB, or None if not found."""
    pid = database.get_project_id(project_name)
    if pid is None:
        return None
    return database.load_project_excel(pid)


def _extract_marca_for_h(marca_raw) -> str:
    """
    Extract the search key from the DB marca field.
    Returns the text AFTER the last '-' in the string (used to search column AM).
    Example: "P62989-07031" → "07031",  "S355-J0" → "J0",  "S355J2" → "S355J2"
    """
    if not marca_raw:
        return ""
    s = str(marca_raw).strip()
    if "-" in s:
        return s.rsplit("-", 1)[1].strip()
    return s


def _get_existing_lines(ws) -> set:
    """
    Return set of (paquete_code, marca, piezas) tuples from data rows (14+).
    Reads columns A, H, I — all plain values written by us, never formulas.
    """
    existing = set()
    for row_num in range(14, ws.max_row + 1):
        col_a = ws.cell(row=row_num, column=1).value
        if col_a:
            col_h = ws.cell(row=row_num, column=8).value
            col_i = ws.cell(row=row_num, column=9).value
            existing.add((
                str(col_a),
                str(col_h) if col_h is not None else "",
                str(col_i) if col_i is not None else "",
            ))
    return existing


def _clear_data_cols(ws, row_num: int) -> None:
    """Clear only columns A, B, H, I for the given row. Leaves formulas in C–G intact."""
    for col_idx in _DATA_COLS:
        ws.cell(row=row_num, column=col_idx).value = None


# ── Public API ────────────────────────────────────────────────────────────────

def excel_exists(project_name: str) -> bool:
    pid = database.get_project_id(project_name)
    if pid is None:
        return False
    return database.excel_exists_in_db(pid)


def delete_project_excel(project_name: str) -> bool:
    pid = database.get_project_id(project_name)
    if pid is None:
        return False
    return database.delete_project_excel_from_db(pid)


def read_excel_preview(project_name: str) -> list[dict]:
    excel_bytes = _load_excel_bytes(project_name)
    if not excel_bytes:
        return []
    try:
        wb = load_workbook(BytesIO(excel_bytes))
        ws = wb.active
        result = []
        for row_num in range(12, ws.max_row + 1):
            row = ws[row_num]
            col_b_value = row[1].value
            if row_num <= 13:
                values = tuple(cell.value for cell in row[0:9])
                result.append({"row": row_num, "values": values})
            elif col_b_value:
                values = tuple(cell.value for cell in row[0:9])
                result.append({"row": row_num, "values": values})
        wb.close()
        return result
    except Exception as e:
        raise ValueError(f"Error al leer Excel: {str(e)}")


def find_marca_in_column_am(project_name: str, marca: str) -> str:
    """Search for marca value in column AM and return full matching value."""
    excel_bytes = _load_excel_bytes(project_name)
    if not excel_bytes:
        return ""
    try:
        wb = load_workbook(BytesIO(excel_bytes), data_only=True)
        ws = wb.active
        col_am = 39
        marca_str = str(marca).strip()
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col_am, max_col=col_am):
            cell_value = row[0].value
            if cell_value and marca_str in str(cell_value):
                wb.close()
                return str(cell_value)
        wb.close()
        return ""
    except Exception as e:
        raise ValueError(f"Error al buscar marca en AM: {str(e)}")


def _build_am_cache(excel_bytes: bytes) -> list[str]:
    """
    Read all non-empty values from column AM (col 39) using data_only=True
    so formula-cached values are returned (same behaviour as find_marca_in_column_am).
    Returns a list of stripped strings.
    """
    col_am = 39
    wb = load_workbook(BytesIO(excel_bytes), data_only=True)
    ws = wb.active
    values = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col_am, max_col=col_am):
        v = row[0].value
        if v is not None:
            values.append(str(v).strip())
    wb.close()
    return values


def _lookup_marca_in_am(marca_raw: str, am_values: list[str]) -> str:
    """
    Find the matching value in the AM column for a given DB marca.

    The DB marca field (e.g. "07031") is used as search key.
    An AM value matches if it ends with that key (e.g. "00010 / P62989-07031").

    Search strategy (in order):
      1. AM value ends with the marca key (e.g. endswith "07031").
      2. AM value contains the marca key as substring (looser fallback).
      3. Fallback: return the marca key itself.
    """
    if not marca_raw:
        return ""
    key = str(marca_raw).strip()

    for v in am_values:
        if v.endswith(key):
            return v
    for v in am_values:
        if key in v:
            return v
    return key  # fallback: write the raw marca if nothing found in AM


def write_lines_to_excel(project_name: str, lines: list[dict]) -> dict:
    """
    Write project lines to Excel starting at row 14. Each DB line → one Excel row.
    Multiple lines per paquete are supported (different marca / piezas).

    Columns written per row:
      A  — paquete_code  (B5/NNNN — prefix read from cell B5, fallback to project_name)
      B  — "Bundle"
      H  — value found in column AM that matches the DB marca field
      I  — piezas

    Col G formula preserved (never touched). Cols C–F left untouched.

    Duplicate key: (paquete_code, marca_h, piezas_str) — stable across syncs
    because marca_h is read back from col H (plain value written by us).

    Saves updated Excel back to database. Returns dict: {added, duplicates, errors}
    """
    pid = database.get_project_id(project_name)
    if pid is None:
        raise ValueError("Proyecto no encontrado.")
    excel_bytes = database.load_project_excel(pid)
    if not excel_bytes:
        raise ValueError("Sube un Excel base primero")

    # Build AM lookup cache using data_only=True (gets formula-cached values)
    am_values = _build_am_cache(excel_bytes)

    # Load for writing (preserves formulas in G etc.)
    wb = load_workbook(BytesIO(excel_bytes))
    ws = wb.active

    # Read the prefix for column A from cell B5 (fallback to project_name if empty)
    b5_value = ws.cell(row=5, column=2).value
    col_a_prefix = str(b5_value).strip() if b5_value else project_name

    existing_lines = _get_existing_lines(ws)

    added = 0
    duplicates = 0
    errors = []

    # Find first row where col A is empty, starting from row 14
    current_row = 14
    for row_num in range(14, ws.max_row + 1):
        if ws.cell(row=row_num, column=1).value is None:
            current_row = row_num
            break
    else:
        current_row = ws.max_row + 1

    for line in lines:
        try:
            paquete_num = line.get("paquete_num")
            paquete_num_str = (
                f"{paquete_num:04d}" if isinstance(paquete_num, int)
                else str(paquete_num).zfill(4)
            )
            paquete_code = f"{col_a_prefix}/{paquete_num_str}"

            # H: value from column AM that best matches the DB marca
            marca_h = _lookup_marca_in_am(line.get("marca", ""), am_values)
            piezas   = line.get("piezas", "")
            piezas_str = str(piezas) if piezas else ""

            line_tuple = (paquete_code, marca_h, piezas_str)
            if line_tuple in existing_lines:
                duplicates += 1
                continue

            # Col A: paquete code  |  Col B: tipo fijo
            # Col G (7): fórmula ya en el template — no se toca
            # Col H (8): valor encontrado en columna AM
            # Col I (9): piezas
            ws.cell(row=current_row, column=1).value = paquete_code
            ws.cell(row=current_row, column=2).value = "Bundle"
            ws.cell(row=current_row, column=8).value = marca_h
            ws.cell(row=current_row, column=9).value = piezas if piezas else ""

            added += 1
            current_row += 1
            existing_lines.add(line_tuple)

        except Exception as e:
            errors.append(f"Fila {current_row}: {str(e)}")

    buf = BytesIO()
    wb.save(buf)
    wb.close()
    database.save_project_excel(pid, buf.getvalue())

    return {"added": added, "duplicates": duplicates, "errors": errors}


def count_n_pedido_rows_in_excel(project_name: str, n_pedido: str) -> int:
    """Count Excel rows that belong to a specific N.Pedido."""
    if not excel_exists(project_name):
        return 0
    pid = database.get_project_id(project_name)
    if pid is None:
        return 0

    lines_db = database.get_paquete_nums_for_n_pedido(pid, n_pedido)
    if not lines_db:
        return 0

    excel_bytes = database.load_project_excel(pid)
    if not excel_bytes:
        return 0

    paquete_codes = set()
    for l in lines_db:
        pnum = l["paquete_num"]
        pnum_str = f"{pnum:04d}" if isinstance(pnum, int) else str(pnum).zfill(4)
        paquete_codes.add(f"{project_name}/{pnum_str}")

    try:
        wb = load_workbook(BytesIO(excel_bytes))
        ws = wb.active
        count = sum(
            1 for row_num in range(14, ws.max_row + 1)
            if ws.cell(row=row_num, column=1).value
            and str(ws.cell(row=row_num, column=1).value) in paquete_codes
        )
        wb.close()
        return count
    except Exception:
        return 0


def delete_all_data_rows_from_excel(project_name: str) -> int:
    """
    Clear data columns (A, B, H, I) for all rows from row 14 onwards.
    Formulas in columns C–G are preserved.
    Saves updated Excel back to database. Returns count of rows cleared.
    """
    pid = database.get_project_id(project_name)
    if pid is None:
        return 0
    excel_bytes = database.load_project_excel(pid)
    if not excel_bytes:
        return 0
    try:
        wb = load_workbook(BytesIO(excel_bytes))
        ws = wb.active
        cleared = 0
        for row_num in range(14, ws.max_row + 1):
            if ws.cell(row=row_num, column=1).value is not None:
                _clear_data_cols(ws, row_num)
                cleared += 1
        buf = BytesIO()
        wb.save(buf)
        wb.close()
        database.save_project_excel(pid, buf.getvalue())
        return cleared
    except Exception as e:
        raise ValueError(f"Error al borrar todas las filas: {str(e)}")


def delete_rows_from_excel(project_name: str, row_numbers: list[int]) -> int:
    """
    Clear data columns (A, B, H, I) for the specified rows (1-indexed).
    Only rows >= 14 are processed (header rows 12-13 are protected).
    Formulas in columns C–G are preserved.
    Saves updated Excel back to database. Returns count of rows cleared.
    """
    safe_rows = [r for r in row_numbers if r >= 14]
    if not safe_rows:
        return 0
    pid = database.get_project_id(project_name)
    if pid is None:
        return 0
    excel_bytes = database.load_project_excel(pid)
    if not excel_bytes:
        return 0
    try:
        wb = load_workbook(BytesIO(excel_bytes))
        ws = wb.active
        for row_num in safe_rows:
            _clear_data_cols(ws, row_num)
        buf = BytesIO()
        wb.save(buf)
        wb.close()
        database.save_project_excel(pid, buf.getvalue())
        return len(safe_rows)
    except Exception as e:
        raise ValueError(f"Error al borrar filas del Excel: {str(e)}")


def delete_n_pedido_from_excel(project_name: str, n_pedido: str) -> int:
    """
    Clear data columns (A, B, H, I) for all rows belonging to a specific N.Pedido.
    Formulas in columns C–G are preserved.
    Saves updated Excel back to database. Returns count of rows cleared.
    """
    if not excel_exists(project_name):
        return 0
    pid = database.get_project_id(project_name)
    if pid is None:
        return 0

    lines_db = database.get_paquete_nums_for_n_pedido(pid, n_pedido)
    if not lines_db:
        return 0

    excel_bytes = database.load_project_excel(pid)
    if not excel_bytes:
        return 0

    paquete_codes = set()
    for l in lines_db:
        pnum = l["paquete_num"]
        pnum_str = f"{pnum:04d}" if isinstance(pnum, int) else str(pnum).zfill(4)
        paquete_codes.add(f"{project_name}/{pnum_str}")

    try:
        wb = load_workbook(BytesIO(excel_bytes))
        ws = wb.active
        cleared = 0
        for row_num in range(14, ws.max_row + 1):
            col_a = ws.cell(row=row_num, column=1).value
            if col_a and str(col_a) in paquete_codes:
                _clear_data_cols(ws, row_num)
                cleared += 1
        buf = BytesIO()
        wb.save(buf)
        wb.close()
        database.save_project_excel(pid, buf.getvalue())
        return cleared
    except Exception as e:
        raise ValueError(f"Error al borrar N.Pedido del Excel: {str(e)}")
